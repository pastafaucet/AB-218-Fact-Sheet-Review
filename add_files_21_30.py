import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Read the existing 20 files
existing_df = pd.read_excel('Enhanced_First_20_Files_VERBATIM_COMPLETE.xlsx')

# Create the data for files 21-30 with VERBATIM text extraction
files_21_30_data = [
    # File 21: Richard Helm
    {
        'Plaintiff Full Name': 'Richard Helm',
        'Completeness': '''Mostly Complete (70%)

✅ Personal Information: Partially complete (father name only, no DOBs)
❌ Education History: NONE PROVIDED
❌ Employment History: NONE PROVIDED
✅ Probation History: Complete (2 facilities, 2008-2010)
⚠️ Perpetrator Identification: Partially complete (about 3 individuals, Ms. Kim identified)
✅ Abuse Details: Detailed with specific conduct
❌ Knowledge/Complaints: No complaints made
✅ Damages/Injuries: Complete (generic template language)
✅ Medical Treatment: Complete
✅ Loss of Earnings: Claims made, amounts TBD
✅ Other Damages: Complete (generic template language)
✅ Documents: Employment and medical records available

NOTE: Verification date 08/18/2025 - signature appears to be "Richard Helm" (consistent with plaintiff name, unlike file 20)''',
        'Perpetrator Details': '''CAMP CHALLENGER (2008-2010, Age 16-17)
Number: About 3 individuals (cannot recall specific number, recalls overall male staff)
Perp 1: Male, Tall
Perp 2: Male, Tall
Position: Correctional officers

CENTRAL JUVENILE HALL (4 different times, Ages 14, 15, 16, 17)
Ms. Kim: Asian Woman, Short, black hair
Position: Staff member''',
        'Summary of Abuse': 'Strip searches with digital penetration, forced masturbation, voyeurism',
        'Abuse Details': '''CAMP CHALLENGER (2008-2010, Age 16-17, Multiple times a week)

Perpetrators conducted constant unnecessary strip searches that included prolonged cavity searches, sexual comments, unnecessary groping of the genitals and buttocks and digital penetration. Plaintiff recalls that digital penetration occurred 6 times. These "strip searches" were done as punishment for fighting.

Location: Hole and plaintiff's room

CENTRAL JUVENILE HALL (At least once a week)

MS. KIM: Force him to masturbate in front of her, watched him in the shower as well.

Location: Not specified''',
        'Mental Health Treatment': '''Dr. Neda Khodaparest, The Green Room Psychological Services
Dates: Various dates
Treatment: Therapy to treat PTSD and other diagnoses

Address: 5252 Balboa Ave., Suite 502, San Diego, CA 92117
Phone: 858-480-9118''',
        'Credibility Notes': '''Moderate credibility concerns: Two facilities with distinct abuse patterns. Camp Challenger: Multiple male staff (about 3, no names, minimal descriptions "tall"), constant strip searches, groping, digital penetration 6 times, done as "punishment for fighting." Central: Ms. Kim (Asian woman, short, black hair) forced masturbation in front of her, watched in shower, at least once a week. No education or employment history provided. No complaints made. Generic damages language. Father name provided (Don Helm) but no mother info or DOBs. Limited perpetrator identification (general descriptions only). Abuse documented as frequent and invasive. Some specific details (6 digital penetrations, as punishment, weekly frequency).'''
    },
    
    # File 22: Daniel Thomas Flores
    {
        'Plaintiff Full Name': 'Daniel Thomas Flores',
        'Completeness': '''⚠️ Extremely Sparse (30%)

⚠️ Personal Information: Partially complete (father name only, no DOBs)
❌ Education History: NONE PROVIDED
❌ Employment History: NONE PROVIDED
✅ Probation History: Minimal (1 facility, 1 year)
⚠️ Perpetrator Identification: Vague ("multiple individuals", "general staff")
⚠️ Abuse Details: Very limited and vague
❌ Knowledge/Complaints: No complaints made
✅ Damages/Injuries: Complete (generic template language)
✅ Medical Treatment: Complete
✅ Loss of Earnings: Claims made, amounts TBD
✅ Other Damages: Complete (generic template language)
✅ Documents: Employment and medical records available''',
        'Perpetrator Details': '''CAMP CHALLENGER (2007-2008, Age 17)
Number: Multiple individuals
Description: General Staff (no individual descriptions provided)
Position: General Staff''',
        'Summary of Abuse': '⚠️ Pepper spray to genitals, physical abuse',
        'Abuse Details': '''CAMP CHALLENGER (2007-2008, Age 17, Multiple Times)

⚠️ VAGUE DESCRIPTION: Plaintiff got into fights with the staff and other inmates, which resulted in front tooth broken. He also got grabbed hard by the staff and throw into the walls. As punishment he would get strip searched, and they would pepper spray his private parts.

Location: Camp Challenger (no specific location within facility provided)''',
        'Mental Health Treatment': '''Dr. Neda Khodaparest, The Green Room Psychological Services
Dates: Various dates
Treatment: Therapy to treat PTSD and other diagnoses

Address: 5252 Balboa Ave., Suite 502, San Diego, CA 92117
Phone: 858-480-9118''',
        'Credibility Notes': '''⚠️⚠️ SIGNIFICANT CREDIBILITY CONCERNS: Extremely vague factsheet. "Multiple individuals" with zero identification or description. Claims pepper spray to genitals as punishment but no specific incidents described. Also claims physical abuse (tooth broken, thrown into walls) mixed with sexual abuse claim. No education history. No employment history. No perpetrator names or descriptions. "Multiple times" without specific frequency. Generic template damages language. Father name provided (Thomas Mario Flores) but no mother info or DOBs. Verification dated 08/15/2025. Unusual abuse allegation (pepper spray to private parts) without substantiating detail. May indicate incomplete intake or difficulty providing specifics.'''
    },
    
    # File 23: Esteban Briano
    {
        'Plaintiff Full Name': 'Esteban Briano',
        'Completeness': '''Mostly Complete (85%)

✅ Personal Information: Complete
✅ Education History: Complete (10th grade)
❌ Employment History: CURRENTLY INCARCERATED (Wasco State Prison, out date 2034, Attempted Murder)
✅ Probation History: Complete (4 facilities, ages 12-17)
✅ Perpetrator Identification: Complete (Michelle DelReal, detailed description)
✅ Abuse Details: Very detailed with progression
✅ Knowledge/Complaints: Told mother and two friends; mother was going to report but he said he would deny it
✅ Damages/Injuries: Detailed with specific additions (distrust of women, current incarceration linked to trauma)
✅ Medical Treatment: Complete (including current Wasco State Prison mental health)
✅ Loss of Earnings: Claims made, amounts TBD
✅ Other Damages: Detailed with specific career impact
✅ Documents: Employment and medical records available''',
        'Perpetrator Details': '''CAMP MUNZ (2011, Age 16)
Number: 1 perpetrator
Name: Michelle DelReal
Description: Female, Salvadorian, Tall, Heavy set, Black Hair, Eyebrows were thick
Position: Probational Officer

First met plaintiff when he was 13 at different facility, remembered him when he returned at 16.''',
        'Summary of Abuse': 'Forced oral sex (giving and receiving), forced vaginal intercourse, grooming, bribery',
        'Abuse Details': '''CAMP MUNZ (2011, Age 16, 3x a week for 180 days)

Job was available as a quarter master. When he got the job the perpetrator would try and wrestle him. She pinned him on the matt and her breasts were in his face and she said she could lose her job so he needs to keep it on low, and she bribed him to keep quiet she would lend him her cellphone and laptop. She made herself the probational officer to client and so she would call him back there whenever and pretend to do "inventory."

One day, she said she would stay an extra day, and she will wear her regular clothes and when she came back there with her regular clothes she would be with client and would give him oral sex, and she forced him to have sex with her one time. She would force client to perform oral sex on her as well. This was a regular basis until he left.

She sat next to him in his cell and gave him her number and gave him 500$ and said she would be there for him, and he wouldn't have to ask his mom for anything.

Location: Laundry Room / Quarter Master

COMPLAINT MADE: Told mother and two friends at Camp Munz. Mother was going to report her but he said he would deny it.''',
        'Mental Health Treatment': '''Dr. Neda Khodaparest, The Green Room Psychological Services
Dates: Various dates
Treatment: Therapy to treat PTSD and other diagnoses

Address: 5252 Balboa Ave., Suite 502, San Diego, CA 92117
Phone: 858-480-9118

Wasco State Prison Mental Health Classes
Dates: 2025 - present
Treatment: Sexual Abuse classes''',
        'Credibility Notes': '''Minor credibility concerns: Single female perpetrator (Michelle DelReal) with detailed identification. Grooming documented (wrestling, breasts in face, threats about losing job, bribery with cellphone/laptop/$500). Manipulation via assignment as his probation officer to gain access. Progression from inappropriate contact to oral sex (giving and receiving) to forced vaginal intercourse. Regular pattern (3x/week for 180 days = ~77 incidents). Contact after release (gave number, $500). Reported to mother and two friends, mother wanted to report but plaintiff said he would deny it. Currently incarcerated at Wasco State Prison (out date 2034) for Attempted Murder. Specific trauma documented: does not trust women till this day, emotions up and down, does not trust authority figures, severe PTSD/Anxiety/Depression, self-medicated with drugs which led to crime cycle, incarceration possibly linked to sexual abuse trauma. Currently attending sexual abuse classes in prison. Wanted to help youth but incarceration ruined opportunity. Well-documented grooming and manipulation pattern with specific details and corroborating disclosure to family/friends.'''
    },
    
    # File 24: Mark Eugene Ferrel
    {
        'Plaintiff Full Name': 'Mark Eugene Ferrel',
        'Completeness': '''Mostly Complete (80%)

✅ Personal Information: Complete
✅ Education History: Partially complete (7th grade partial, Harbor Occupational Center)
✅ Employment History: Complete
✅ Probation History: Complete (6 facilities, age 13)
✅ Perpetrator Identification: Detailed (Ms. Lanz and Mr. Wolf with descriptions)
✅ Abuse Details: Very detailed with specific progression and witness to additional abuse
✅ Knowledge/Complaints: No complaints made
✅ Damages/Injuries: Detailed (specific impacts: joined gang, ran away, stealing cars, reckless behavior, fights, crime cycle, prison)
✅ Medical Treatment: Complete
✅ Loss of Earnings: Claims made, amounts TBD
✅ Other Damages: Detailed with specific career impact (wanted to enlist in Navy like father, blocked by felony)
✅ Documents: Employment and medical records available
✅ Prior Litigation: Yes (2017, 24 hour fitness wrongful termination, settled)''',
        'Perpetrator Details': '''CAMP MILLER (1990, Age 13)
Number: 2 perpetrators

Perpetrator 1 (Ms. Lanz): Female, African American - light skinned or could be mixed/Puerto rican - really light skinned, Tall, Pigmentation/Brown on her face/ Freckles, Big breasts, Thin in her leg area, Case Worker

Perpetrator 2 (Mr. Wolf): Male, Caucasian, Older, 5'6-5'7, Rounded, Santa Clause looking - grey hair with grey beard, Very distinctive voice - Deep, Director''',
        'Summary of Abuse': 'Grooming, forced masturbation, witnessed oral sex on another child, threats',
        'Abuse Details': '''CAMP MILLER (1990, Age 13, P1: 4-5x, P2: witnessed 2x)

PERPETRATOR 1 (MS. LANZ) - 4-5x:
She introduced herself and stated she would be his case worker. Merit Ladder system: top 10 or 2-3 get special privileges (leave camp, take walks, eat popcorn, watch movies, coupons for days off). She held client back and told him that he needs to be #1 on the top 5 on the Merit Ladder and if he fails under that then she will give him an added day coupon which results in trouble.

Next day - invited top 10 to movie on Friday night in office. Perpetrator told client to go into the office with her and had an adjoined drawer next to it. When she brought him in there she forced client to kiss her and made him touch her breasts. She told client not to discuss this with anyone, no one in camp or his parents and also asked him about his parents. She said if they come and visit and he tells anyone she will make sure to write him up so that he will be locked up in there for a very long time.

Client was groomed and was told to maintain #1 on the Merit Ladder no matter what. This started to become the norm where the perpetrator would come every day or every other day. She made client walk in the back with her and Perpetrator 2; Mr. Wolf was in the front. She told Mr Wolf to go up and she stood back with client and pulled client to the side of the bin and started groping him and masturbated him off and he states he never done this before and did not know what it was, she made him ejaculate from it. She walked them back up to where everyone else was.

FIELD TRIP INCIDENT:
A week later - she told client they were going on a field trip and again that he needed to be #1 on the Merit Ladder so he could go on the field trip with them. After truck was packed, they left and was somewhere in the mountains, there were white signs with numbers on them (mile/elevation markers). They went up to a cabin around 700 mile/elevation marker - they unpacked and Ms. Lanz was not there so client felt more at ease. Mr. Wolf and one child was in one room and two other rooms were 4 kids each on the other rooms.

PERPETRATOR 2 (MR. WOLF) - WITNESSED 2x:
When client accidently opened the wrong door that was supposed to be the bathroom - he saw Mr. Wolf performing Fellatio on one of the clients. Mr. Wolf came out and grabbed client's head and pulled him to the other room and told him not to say anything to anyone and threw him against the wall and he kept promising him he wasn't going to say anything. Mr. Wolf forced everyone outside where the firewood was being cut and pretended nothing happened and client was frozen up and did not hear much of anything afterwards and the camp fieldtrip was ended early.

Location: P1: Office and Outside. P2: Camp field trip / Cabin

THREATS: Ms. Lanz: threatened extended detention if he told anyone. Mr. Wolf: grabbed his head, threw him against wall, threatened him not to tell after he witnessed abuse of another child.''',
        'Mental Health Treatment': '''Dr. Neda Khodaparest, The Green Room Psychological Services
Dates: Various dates
Treatment: Therapy to treat PTSD and other diagnoses

Address: 5252 Balboa Ave., Suite 502, San Diego, CA 92117
Phone: 858-480-9118''',
        'Credibility Notes': '''Minor credibility concerns: Two perpetrators with detailed accounts. P1 (Ms. Lanz): Female case worker, grooming via merit ladder system, forced kissing/breast touching, forced masturbation, threats of extended detention. Detailed physical description (light-skinned African American or mixed, freckles, big breasts). P2 (Mr. Wolf): Director, distinctive appearance (Santa Clause looking, grey hair/beard, deep voice). Client witnessed Mr. Wolf performing oral sex on another child during field trip. Physical violence when discovered (grabbed head, threw against wall). Field trip ended early after incident. Specific details: cabin at 700 mile/elevation marker, merit ladder system mechanics, Friday night movie privileges. Post-abuse impacts well documented: extremely shy before, joined gang after, ran away from home, stole parents' cars, lashed out, reckless/ruthless behavior, many fights, turned to crime, went to prison for robbery, violated parole, cycle continued long time, severe PTSD/anxiety/depression, constant nightmares, trust issues with authority. Wanted to enlist in Navy like father but felony at 18 blocked this path. Prior litigation (2017 wrongful termination case, settled). Detailed grooming pattern, witness to abuse of others, specific threats and violence.'''
    },
    
    # File 25: Teona Hunter
    {
        'Plaintiff Full Name': 'Teona Hunter',
        'Completeness': '''Mostly Complete (75%)

✅ Personal Information: Complete
✅ Education History: Complete (12th grade diploma from Job Core)
❌ Employment History: NONE in last 5 years
✅ Probation History: Complete (1 facility, 2006-2009)
⚠️ Perpetrator Identification: Partially complete (detailed description but no name)
✅ Abuse Details: Detailed with specific incidents
✅ Knowledge/Complaints: Told Staff Member Rachel who said client was "exaggerating"; perpetrator then assigned to her more; eventually moved to different unit
✅ Damages/Injuries: Complete
✅ Medical Treatment: Complete
✅ Loss of Earnings: Claims made, amounts TBD
✅ Other Damages: Complete
✅ Documents: Employment and medical records available''',
        'Perpetrator Details': '''CAMP MCNAIR (2006-2009, Age 13-16)
Number: 1 perpetrator
Description: Male, Caucasian, Age range: 30's, receding hairline, blue or brown eyes, 5'7, heavyset - beer gut, Scar on his left arm in between elbow and forearm 4-5inches and keloid
Position: Staff Member''',
        'Summary of Abuse': 'Groping, attempted digital penetration, voyeurism (masturbating during strip search)',
        'Abuse Details': '''CAMP MCNAIR (2006-2009, Age 13-16, 4-5x until moved to a different unit)

Client noted that she was groped by a CO. She mentioned that a guard would come into her cell and "check if she was alive." She also mentioned that she was stripped down naked he would attempt to help her take her clothes off and told her, "You're moving to slow". Once she got naked, he would fondle her breasts. When she turned around and bent over, he attempted to penetrate her with her fingers and then she kicked him.

One other incident she noticed he was jacking off during another strip search.

He would take her calls away if she was refusing.

Location: Intake

COMPLAINT MADE: Told Staff Member: Rachel. The staff member told her that the client was exaggerating. When she kept telling about the sexual assault, the perpetrator was assigned to her more. Eventually she got moved to a different unit.''',
        'Mental Health Treatment': '''Dr. Neda Khodaparest, The Green Room Psychological Services
Dates: Various dates
Treatment: Therapy to treat PTSD and other diagnoses

Address: 5252 Balboa Ave., Suite 502, San Diego, CA 92117
Phone: 858-480-9118''',
        'Credibility Notes': '''Moderate credibility concerns: Single male perpetrator (Caucasian, 30s, receding hairline, heavyset with beer gut, distinctive scar on left arm 4-5 inches with keloid). Detailed description but no name. Abuse pattern: came to cell to "check if she was alive," forced strip searches, told her "You're moving to slow" when she undressed, fondled breasts, attempted digital penetration when she bent over (she kicked him in response), observed masturbating during another strip search. Coercion: would take away phone calls if she refused. Reported to Staff Member named Rachel who dismissed claim saying client was "exaggerating." Retaliation documented: after reporting, perpetrator was assigned to her MORE. Eventually moved to different unit. Female victim. No employment in last 5 years. Born in Dayton, OH. Completed high school via Job Core in Cleveland, OH. Three-year detention period (2006-2009, ages 13-16). Direct quote from perpetrator. Responsive behavior documented (kicked him). Failed reporting with retaliation. Specific physical identifier (distinctive scar). Reason for probation not provided.'''
    },
    
    # File 26: Byron Ramon Lopez Jr.
    {
        'Plaintiff Full Name': 'Byron Ramon Lopez Jr.',
        'Completeness': '''⚠️ Critically Incomplete (20%)

❌ Personal Information: Almost entirely blank (no parent names or DOBs)
❌ Education History: COMPLETELY BLANK
❌ Employment History: NONE PROVIDED
✅ Probation History: Minimal (1 facility, in and out 9 months total)
⚠️ Perpetrator Identification: Vague (1 individual, minimal description)
⚠️ Abuse Details: Limited detail, vague frequency
❌ Knowledge/Complaints: No complaints made
✅ Damages/Injuries: Complete (generic template language)
✅ Medical Treatment: Complete
✅ Loss of Earnings: Claims made, amounts TBD
✅ Other Damages: Partially complete
✅ Documents: Employment and medical records available''',
        'Perpetrator Details': '''LOS PADRINOS (2013, Age 14)
Number: 1 individual
Description: Male, African American, medium build, probably late 20s early 30s, short curly hair
Position: PO (Probation Officer)''',
        'Summary of Abuse': 'Strip searches, genital touching, sexual comments, threats',
        'Abuse Details': '''LOS PADRINOS (2013, Age 14, multiple times every day for 2 weeks)

Constantly every day, the plaintiff was pulled aside individually and strip searched. He was singled out after school and taken alone to the bathroom. On the perpetrator's days off, no other staff would strip search the plaintiff.

The perpetrator forced the plaintiff to take off his clothes completely and touched the plaintiff's genitals and waist. One time while the plaintiff was putting his clothes back on, the perpetrator made a sexual comment about the plaintiff's genitals asking if it "grows."

The perpetrator threatened the plaintiff with being forced to stay in his cell if he did not comply with the searches. He was also threatened with losing privileges like watching TV and recreation. He was sometimes bribed with small snacks.

Location: Plaintiff's cell''',
        'Mental Health Treatment': '''Dr. Neda Khodaparest, The Green Room Psychological Services
Dates: Various dates
Treatment: Therapy to treat PTSD and other diagnoses

Address: 5252 Balboa Ave., Suite 502, San Diego, CA 92117
Phone: 858-480-9118''',
        'Credibility Notes': '''⚠️⚠️ SIGNIFICANT CREDIBILITY CONCERNS: Factsheet is nearly blank. NO education history provided at all. NO employment history. NO personal information (no parent names or DOBs). Single perpetrator (African American male PO, late 20s/early 30s, medium build, short curly hair) with minimal description. Abuse description is vague: "constantly every day" strip searches for 2 weeks, genital touching, sexual comment asking if genitals "grows," threats (forced to stay in cell, lose privileges), bribes (small snacks). Singled out after school, taken to bathroom alone, only happened on this perpetrator's days. Generic template damages language. Verification dated 04/23/2025. Pattern suggests incomplete intake or minimal cooperation in providing details. No specific perpetrator name. Two-week timeframe unusual (most cases span months/years). Educational/career impact noted but no baseline education provided. Reason for probation: Battery and vandalism.'''
    },
    
    # File 27: Renard Reynolds
    {
        'Plaintiff Full Name': 'Renard Reynolds',
        'Completeness': '''Mostly Complete (75%)

✅ Personal Information: Complete (father not known)
✅ Education History: Complete (high school diploma, some college)
✅ Employment History: Retired
✅ Probation History: Minimal (1 facility, 1965-1966, age 6 or 7)
⚠️ Perpetrator Identification: Very vague (1 perpetrator, minimal description)
✅ Abuse Details: Complete with specific acts
✅ Knowledge/Complaints: Verbal complaint to Ms. Dollar (White female) describing sexual assault, but no response
✅ Damages/Injuries: Detailed with specific examples
✅ Medical Treatment: Complete
✅ Loss of Earnings: Claims made, amounts TBD
✅ Other Damages: Detailed with specific career impact
✅ Documents: Employment and medical records available

⚠️ CRITICAL NOTE: Plaintiff was age 6 or 7 at time of abuse (1965-1966) - EXTREMELY YOUNG for probation placement. Father's discretion listed as reason.''',
        'Perpetrator Details': '''CAMP CHALLENGER (1965-1966, Age 6 or 7)
Number: 1 perpetrator
Description: Male, White (minimal description provided)
Position: Counselor

Perpetrator promised plaintiff to help find plaintiff's lost brothers.''',
        'Summary of Abuse': 'Genital touching, forced masturbation (giving and receiving), forced oral sex (receiving)',
        'Abuse Details': '''CAMP CHALLENGER (1965-1966, Age 6 or 7, twice)

The perpetrator touched the plaintiff's genitals under clothing. The perpetrator masturbated the plaintiff and forced the plaintiff to masturbate the perpetrator. The perpetrator forced the plaintiff to receive oral copulation.

Location: The perpetrator's office

COMPLAINT MADE: Plaintiff made verbal complaint to Ms. Dollar (White female) describing the sexual assault that occurred. No response documented.''',
        'Mental Health Treatment': '''Dr. Neda Khodaparest, The Green Room Psychological Services
Dates: Various dates
Treatment: Therapy to treat PTSD and other diagnoses

Address: 5252 Balboa Ave., Suite 502, San Diego, CA 92117
Phone: 858-480-9118''',
        'Credibility Notes': '''⚠️ UNUSUAL CASE - EXTREMELY YOUNG VICTIM: Plaintiff was only 6 or 7 years old at time of abuse in 1965-1966. Reason for probation: "Father's discretion" - highly unusual. Perpetrator groomed child by promising to help find lost brothers. Single perpetrator (male, White, counselor) with minimal description. Specific acts documented: genital touching under clothing, forced mutual masturbation, forced to receive oral copulation. Occurred twice in perpetrator's office. Complaint made to Ms. Dollar (White female) but no response. Now 67 years old (DOB 10/27/1957), currently retired. Completed high school (Monrovia High School), attended Citrus College at age 18 (incomplete). Specific impacts documented: self-blame, low self-esteem, depression, anxiety, reincarceration, emotional deregulation, difficulty connecting with friends/family, difficulty maintaining romantic relationships, issues with sexual intimacy. Changed career goals from law enforcement due to abuse. Age at abuse (6-7) is extraordinarily young for juvenile detention system, raising questions about placement circumstances. 60-year delay in reporting. Despite young age at abuse, specific acts remembered.'''
    },
    
    # File 28: Steven Pilcher
    {
        'Plaintiff Full Name': 'Steven Pilcher',
        'Completeness': '''Mostly Complete (80%)

✅ Personal Information: Complete
✅ Education History: Complete (11th grade)
❌ Employment History: Currently unemployed
✅ Probation History: Complete (4 facilities, ages 13-15)
⚠️ Perpetrator Identification: Partially complete (3 perpetrators, descriptions but no names)
✅ Abuse Details: Very detailed with three separate perpetrators and specific acts
❌ Knowledge/Complaints: No complaints made
✅ Damages/Injuries: Detailed
✅ Medical Treatment: Complete (includes Kaiser Alcott Center for Mental Health since 1998)
✅ Loss of Earnings: Claims made, amounts TBD
✅ Other Damages: Complete
✅ Documents: Employment and medical records available''',
        'Perpetrator Details': '''CAMP CHALLENGER (1996/97, Age 14/15)
Number: 3 perpetrators

P1: Male, Hispanic, mid 30's, clean cut, Heavyset, clean-cut haircut, Staff Member

P2: Male, Hispanic, dark hair - combed hair, mustache, medium size/healthy, Staff Member  

P3: Male, Caucasian, regular haircut/fade, Tall, Strong/Muscular build, hazel eyes, Staff Member''',
        'Summary of Abuse': 'Forced kissing, groping, anal penetration, forced oral sex (giving and receiving)',
        'Abuse Details': '''CAMP CHALLENGER (1996/97, Age 14/15, P1: 4x, P2: 4x, P3: 5x)

PERPETRATOR 1 - 4x:
Perpetrator would force the plaintiff to kiss him. He would also grope/fondle the plaintiff's genitals and penile penetrate him in closet areas.

PERPETRATOR 2 - 4x:
The first time, the perpetrator made client perform oral copulation on him. Second incident perpetrator walked client to the gym area and when the doors opened he made client stand in a position up against the wall and told client to pull his pants down and he started to fondle client and would compliment him on his smell and then he would kiss client and then told client to give him oral sex and telling him how he tasted. Other incidents with the perpetrator did the same thing happened that included fondling him and continuing to give him oral sex.

PERPETRATOR 3 - 5x:
Perpetrator started fondling client and started giving oral sex to client until he ejaculated. The following incidents were of a similar manner where client was sat/brushed on top of a table and was told to pull his pants down and then perpetrator forced client to give him oral sex and vice versa.

Location: Closet, Clothing Area, Shower, Bathroom, Office''',
        'Mental Health Treatment': '''Dr. Neda Khodaparest, The Green Room Psychological Services
Dates: Various dates
Treatment: Therapy to treat PTSD and other diagnoses

Address: 5252 Balboa Ave., Suite 502, San Diego, CA 92117
Phone: 858-480-9118

Dr. Kaiser Alcott Center For Mental Health
Dates: 1998 - Present
Treatment: Mental Health''',
        'Credibility Notes': '''Moderate credibility concerns: Three perpetrators at same facility with distinct patterns. P1 (Hispanic, mid-30s, heavyset): forced kissing, groping, anal penetration in closet areas (4x). P2 (Hispanic, mustache): forced oral sex (giving), fondling, kissing, compliments on smell, comments on taste (4x). P3 (Caucasian, tall, muscular, hazel eyes): forced oral sex (giving and receiving), fondling, made client ejaculate (5x). Multiple locations: closet, clothing area, shower, bathroom, office. Total 13 incidents across three perpetrators. Currently unemployed. Has been receiving mental health treatment since 1998 (27 years of documented treatment). Completed education through 11th grade. Detailed account with specific progression for each perpetrator. Impacts documented: nightmares, flashbacks, tormented by abuse, hard to talk about, PTSD/anxiety/depression. Reason for probation: GTA (Grand Theft Auto). No perpetrator names provided despite detailed descriptions. Long history of mental health treatment supports trauma claim.'''
    },
    
    # File 29: Masood Haidari
    {
        'Plaintiff Full Name': 'Masood Haidari',
        'Completeness': '''Mostly Complete (85%)

✅ Personal Information: Complete
✅ Education History: Complete (high school diploma, culinary arts degree)
✅ Employment History: Complete (9 years at current job)
✅ Probation History: Complete (3 facilities, ages 13-17)
⚠️ Perpetrator Identification: Partially complete (multiple perpetrators, some descriptions)
✅ Abuse Details: Very detailed across multiple facilities
✅ Knowledge/Complaints: Complained to older Caucasian lady staff member about fear of being silent
✅ Damages/Injuries: Very detailed with extensive specific impacts including suicide attempt
✅ Medical Treatment: Complete
✅ Loss of Earnings: Claims made, amounts TBD
✅ Other Damages: Very detailed with extensive career/education impact
✅ Documents: Employment and medical records available
✅ Prior Litigation: Yes (2020 against AT&T, gave credit on bill)''',
        'Perpetrator Details': '''LOS PADRINOS & BARRY J. NIDORF (1995-1997, Ages 13-15)
Number: Mostly Males
Description: Some Caucasian, some Hispanic, some African American
Position: Staff members

BARRY J. NIDORF (specific incidents):
P1: Male, Caucasian, Tall
P2: Male, Hispanic, Tall

CAMP CHALLENGER/JARVIS (1997-1999, Ages 15-17)
Male, Hispanic, 5'9, military person with water bottle military style''',
        'Summary of Abuse': 'Sexualized strip searches, voyeurism with masturbation, forced sexualized fighting, groping/fondling, sexual comments, retaliation',
        'Abuse Details': '''LOS PADRINOS & NIDORF (1995-1997, Ages 13-15, Multiple occasions)

Coming out the shower perps would smack plaintiff's butt over the towel, and massage chest in sexual way during sexualized strip searches. When males were there the towels would be ripped, resulting in them walking half-naked. Perps would watch sexually and say sexualized comments.

In the showers there was a section where the perps would sit and masturbate their own genitals while watching the plaintiff shower (voyeurism).

All perps would set up sexualized fights where they forced the plaintiff to engage in sexualized horseplay. In these fights, the plaintiff was forced to only wear boxers.

CAMP CHALLENGER-JARVIS (1997-1999, Ages 15-17, Multiple occasions)

Perp would lay in the bunks of the dorm and would play oldies music. Sexual comments were made by the perp during nighttime.

Perp would conduct this "game" where he would force the plaintiff to cover his eyes with towels and would start to fondle/grope the plaintiff's penile region, chest and bottom. Perp would start by his head, shoulders, thigh area, close to genitals and then genitals over clothing. When the plaintiff did not comply to the "game" he was sent to the shoe for speaking up/refusing.

Voyeurism was also conducted here with sexual comments.

CAMP MALIBU:
Multiple sexualized fights occurred. One staff member would walk around with porn magazines and showed everyone including plaintiff sexual images.

ALL FACILITIES - RETALIATION: If plaintiff did not comply they would take away privileges and send plaintiff to the hole and say things like "no one is trusting you or believing you".

Location: Showers, plaintiff's cell/dorm, the yard

COMPLAINT MADE: Complained to older Caucasian lady staff member that would check the facility about the fear he had in being silent about the sexual abuse.''',
        'Mental Health Treatment': '''Dr. Neda Khodaparest, The Green Room Psychological Services
Dates: Various dates
Treatment: Therapy to treat PTSD and other diagnoses

Address: 5252 Balboa Ave., Suite 502, San Diego, CA 92117
Phone: 858-480-9118''',
        'Credibility Notes': '''Minor credibility concerns: Multiple facilities (Los Padrinos, Barry J. Nidorf, Camp Challenger/Jarvis, Camp Malibu) with different abuse patterns spanning ages 13-17. Multiple perpetrators (Caucasian, Hispanic, African American males). Detailed patterns: sexualized strip searches, butt smacking over towel, chest massage, ripped towels forcing half-naked walking, voyeurism (staff masturbating while watching showers), forced sexualized fights in boxers only, "game" with eyes covered involving progressive groping (head→shoulders→thighs→genitals), nighttime sexual comments with oldies music, pornography shown. Retaliation documented: sent to "the shoe" for refusing, privileges taken, told "no one is trusting you or believing you". Complaint made to older Caucasian lady about fear of being silent. Born in Kabul, Afghanistan. Currently successful (Head Chef for 9 years at Sinful Catering). Culinary arts degree from Lecordonblu at age 33. Extensive detailed impacts: feels world is not safe, deep-rooted distrust, impossible to work under others (forced to create own business), diagnosed anxiety/PTSD/depression, frequent panic attacks, flashbacks, obsessive thoughts tied to abuse, daily reliving, persistent self-blame/low self-worth, constant doubt/second-guessing, trust issues affecting romantic/sexual intimacy, fears commitment, overwhelming paranoia/jealousy, fractured family relationships (partially blames mother), institutionalized feeling, "if one adult had listened lives could have been saved", suicide attempt at 16/17 by taking immense number of pills. Detailed systemic pattern across multiple facilities. Successful career despite trauma. Prior litigation (2020 AT&T). Reason for probation: Truancy.'''
    },
    
    # File 30: Julio Quezada
    {
        'Plaintiff Full Name': 'Julio Quezada',
        'Completeness': '''Mostly Complete (90%)

✅ Personal Information: Complete (father N/A)
✅ Education History: Complete (9th grade, 1 week)
✅ Employment History: Complete (self-employed 5 years)
✅ Probation History: Complete (4 facilities, ages 14-16)
✅ Perpetrator Identification: Very detailed (8 perpetrators across 4 facilities with descriptions, one partial name)
✅ Abuse Details: Extremely detailed with multiple perpetrators and facilities
❌ Knowledge/Complaints: No complaints made
✅ Damages/Injuries: Very detailed with extensive specific impacts
✅ Medical Treatment: Complete
✅ Loss of Earnings: Claims made, amounts TBD
✅ Other Damages: Very detailed with specific impacts
✅ Documents: Employment and medical records available''',
        'Perpetrator Details': '''BARRY J. NIDORF (1994-1996, Ages 14-16)
Number: 4 perpetrators

Perp 1: Male, Hispanic or Caucasian, light skin, heavy set, 300/350 lbs, hair in ponytail, goatee, tattoos on arm maybe on biceps, typical biker look

Perp 2: Male, Hispanic, tall, 6'1/6'2, brown skin, dark complexion, no facial hair, black hair in clean cut maybe combed to the back, seeing glasses, always wore polo shirts from home

Perp 3: Female, African-American, looked mixed, tall, 5'8/5'9, long dark brown hair, slim build, wore a lot of makeup

Perp 4: Female, Hispanic, shorter, a bit on the heavier side, black hair

LOS PADRINOS (1994-1996, Ages 14-16) - Serious offenders' unit
Name: Think last name Durran
Description: Male, Hispanic, short, maybe 5'5/5'6, clean cut hair, strong looking/muscular

CENTRAL JUVENILE HALL (1994-1996, Ages 14-16)
Number: 2 perpetrators

Perp 1: Male, African-American, very light skin, looked mixed with Caucasian, short black clean cut hair, skinny but muscular/toned, tall, maybe 5'9

Perp 2 (Ms. Ortega): Female, Hispanic, short, very chubby, long black hair always in slick back ponytail

CAMP MICHAEL SMITH (1995-1996, Ages 15-16)
Male, Caucasian, older, mid-50s, short maybe 5'5/5'6, blondish grayish long hair in a ponytail, think had light goatee''',
        'Summary of Abuse': 'Digital anal penetration, genital groping, sexual comments/coercion, threats, degradation, attempted rape',
        'Abuse Details': '''BARRY J. NIDORF (1994-1996, Ages 14-16)

PERP 1 (at least 3 physical incidents, sexual comments whenever alone):
Would push himself against plaintiff, attempt to restrain him, grope and slap his bottom, touch his genitals over and under clothing, and degrade him. Digitally anally penetrated plaintiff 2 times. Physical incidents in plaintiff's room while changing clothes, recreational areas, storage rooms like big laundry room. Sexual comments: "I would give anything to penetrate you" and laughing. Anytime caught alone in hallways or recreational area. Threatened: "you better not say anything. No one gonna believe you because you're a criminal."

PERP 2 (3-4 physical incidents):
Sometimes acted like searching plaintiff in his room. Pushed plaintiff against wall, grabbed his bottom more than 3 times and grabbed genital area 1 time, over and under clothing for both areas. Sexual comments saying he wanted to do sexual things include anal rape. Said "I would make you my bitch." Would tell plaintiff to bend over in many different ways and laugh, watching voyeuristically. Threatened: "No one gonna believe you. You're nothing, just a criminal. You're gonna waste your life in and out of the system." "Your word against mine."

PERP 3 (4-5 physical incidents, over 15 incidents of sexual comments):
Made sexual comments telling plaintiff she has not been with a male in years and stating sexual acts he wanted plaintiff to do to her. Grabbed plaintiff's genitals 4-5 times over and under clothing. Told plaintiff it was okay if he touched her sexually. If plaintiff refused, she became mean, degrading him including calling him nothing and saying racist comments. Asked about having intercourse, or female forced/coerced rape. Threatened: "Don't think about it. Don't waste your time. No one will believe you." Would be less harsh if plaintiff agreed to her sexual requests.

PERP 4 (2 physical incidents, at least 10 incidents of sexual comments):
Made sexual comments over 10 times asking about sexual intercourse, or female forced/coerced rape. Grabbed plaintiff's genitals twice, over and under clothing. Threatened as well.

LOS PADRINOS (1994-1996, Ages 14-16, 2-3 physical incidents):
Grabbed plaintiff's genitals and bottom over and under clothing. Did not speak while sexually violating plaintiff. Threatened "Don't say shit" or else he would take plaintiff to the back and "beat his ass." Plaintiff had been physically beat by perpetrator 5-6 times (punched in stomach and back), so took threats seriously.
Location: Dayroom when empty and dark

CENTRAL JUVENILE HALL (1994-1996, Ages 14-16)

PERP 1 (1 incident):
Sexually touched plaintiff's genitals and bottom over and under clothing.
Location: Hallway when exiting dayroom

PERP 2 - MS. ORTEGA (3 physical incidents, sexual comments more than once a day):
Asked plaintiff if he wanted to go in a room with her and do sexual acts to her. Told plaintiff she would not tell anyone. Pushed plaintiff toward wall in laundry room, touched his genitals under clothing, stated she wanted him to do sexual things to her. In one incident, pulled his pants down to his knees. Asked about having sexual intercourse, or female forced/coerced rape. If plaintiff refused, would get angry and degrade him: "Anyone else would be all over this. You're just a little boy." Made sexual comments asking to rape him more than once a day. Comments started before physical touching and lasted until he left facility. Both perpetrators threatened: "Your word doesn't count. You're most likely going to prison."
Location: Laundry room

CAMP MICHAEL SMITH (1995-1996, Ages 15-16, 2 physical incidents, sexual comments twice a day for month and a half):
Made sexual comments twice a day asking plaintiff to have sexual intercourse, or anal rape by male penile penetration. Sexual remarks progressed to physical touching, pushing and shoving plaintiff and touching plaintiff's bottom under clothing. Sexual comments when coming out of shower. Physical touching in bunk bed because no one else was on that row of bunks. Threatened: "Don't think about saying anything. You're wasting your time. No one will believe you." Degraded plaintiff calling him a criminal and saying about sexual abuse: "This is how people like you get treated."
Location: Sexual comments when coming out of shower. Physical touching in bunk bed.''',
        'Mental Health Treatment': '''Dr. Neda Khodaparest, The Green Room Psychological Services
Dates: Various dates
Treatment: Therapy to treat PTSD and other diagnoses

Address: 5252 Balboa Ave., Suite 502, San Diego, CA 92117
Phone: 858-480-9118''',
        'Credibility Notes': '''Minor credibility concerns: Extremely detailed account across 4 facilities with 8 perpetrators spanning ages 14-16. NIDORF had 4 perpetrators (3 with detailed physical descriptions, including "biker look" male who digitally penetrated him twice, tall Hispanic male with glasses who threatened to make him "my bitch," two females who grabbed genitals and made sexual comments/threats). Los Padrinos: One perpetrator (possibly "Durran") who also physically beat him 5-6 times to enforce compliance. Central: Two perpetrators including named Ms. Ortega who pulled his pants down and made rape comments daily. Smith: Older male (mid-50s) with gray ponytail who made comments twice daily for 6 weeks. Multiple direct quotes from perpetrators. Detailed threats and degradation documented. Physical violence to enforce silence. Specific impacts: anxiety, depression, occasional insomnia, night sweats, difficulty connecting with wife/family, distrust of authority, self-medicated with alcohol/marijuana for years, nightmares (constant in teens/20s, occasional now), flashbacks 3x/week, trouble with sexual intimacy, trouble forming/maintaining romantic relationships, trouble eating, heavy alcohol use caused marital trouble (now separated from wife Lisa Ann Quezada), overprotective of kids, tries to block world out, "abuse always creeps back up in happiest moments." Educational impact: trouble paying attention after abuse, lost hope in career goals for years, prefers working independently (self-employed 5 years in auto detail/pressure washing), always felt on edge at jobs reducing productivity. Only attended high school 1 week of 9th grade. Married. Currently self-employed. Original charge: Attempted murder, dropped to assault with deadly weapon. Verification dated 01/30/2025. Extraordinary level of detail across multiple facilities and perpetrators. Consistent pattern of threats ("no one will believe you because you're a criminal"). Multiple female perpetrators documented.'''
    }
]

# Create DataFrame for files 21-30
files_21_30_df = pd.DataFrame(files_21_30_data)

# Combine all 30 files
all_30_df = pd.concat([existing_df, files_21_30_df], ignore_index=True)

# Save to Excel
output_file = 'Enhanced_First_30_Files_VERBATIM_COMPLETE.xlsx'
all_30_df.to_excel(output_file, index=False, engine='openpyxl')

# Format the Excel file
wb = load_workbook(output_file)
ws = wb.active

# Set column widths
ws.column_dimensions['A'].width = 25  # Plaintiff Full Name
ws.column_dimensions['B'].width = 35  # Completeness
ws.column_dimensions['C'].width = 50  # Perpetrator Details
ws.column_dimensions['D'].width = 30  # Summary of Abuse
ws.column_dimensions['E'].width = 80  # Abuse Details
ws.column_dimensions['F'].width = 35  # Mental Health Treatment
ws.column_dimensions['G'].width = 50  # Credibility Notes

# Enable text wrapping for all cells
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=7):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')

wb.save(output_file)

print(f"✅ Successfully created {output_file}")
print(f"   Total entries: {len(all_30_df)}")
print(f"   - Original files 1-20: 20 entries")
print(f"   - New files 21-30: 10 entries")
print(f"\n📊 File breakdown:")
for i, name in enumerate(all_30_df['Plaintiff Full Name'], 1):
    print(f"   {i}. {name}")
