import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Sample data from first 6 files (I'll extract the others)
data_list = [
    # File 1: Rodriguez, Jesse (already extracted)
    {
        'Plaintiff Full Name': 'Jesse Rodriguez',
        'Completeness': '''Mostly Complete (85%)

✅ Personal Information: Complete
✅ Education History: Complete
⚠️ Employment History: Partially Complete
✅ Probation History: Complete (3 facilities)
✅ Perpetrator Identification: Complete
✅ Abuse Details: Complete
❌ Knowledge/Complaints: No complaints made
✅ Damages/Injuries: Complete
✅ Medical Treatment: Complete
⚠️ Loss of Earnings: Claims made, amounts TBD
✅ Other Damages: Complete
⚠️ Documents: Limited''',
        'Perpetrator Details': '''CENTRAL JUVENILE HALL (March 2007)
Number: 1 perpetrator
Description: Male, Hispanic/African American, 39-41 years old, tanned skin, 5'9"-6'0", medium build/in shape, probation officer (evening shift)

BARRY J. NIDORF JUVENILE HALL (2007)
Number: 2 perpetrators
  Perp 1: Male, African American, heavy set (250+ lbs), mustache, dark skin, nasty attitude toward kids, probation officer in charge of unit
  Perp 2: Female, Hispanic/Mexican American, late 30s-40s, tall (6'0"), thick build, verbally abusive

CAMP MICHAEL SMITH (July 2007)
Number: 1 perpetrator
Description: Male, Puerto Rican/Dominican, NY accent, 6'1", short bald haircut, light brown skin, camp counselor''',
        'Abuse Details': '''CENTRAL JUVENILE HALL (March 2007, Age 15-16, 2 incidents)

Perpetrator befriended plaintiff midway through stay, purposely got close, asked questions. Made plaintiff shower last by himself. While plaintiff was naked in shower, perpetrator entered twice. First time claimed to "check", second time brought soap into shower and bathed plaintiff, touching all over plaintiff's body.

After shower, perpetrator stayed in plaintiff's room making sexual remarks and questions. Told plaintiff "I see you as a son" while touching plaintiff's genitals, chest and bottom while plaintiff was naked. Tucked plaintiff into bed saying "you'll have a good life, you're going to grow up handsome" - grooming behavior.

Perpetrator said "you remind me of my kid, I can take care of you" while touching and caressing plaintiff's leg down to foot in soft, caressing tone. Watched plaintiff heavily while showering.

In separate incident, plaintiff woke to perpetrator touching plaintiff's genitals under clothing. When plaintiff woke and asked what he was doing, perpetrator shrugged it off, laughed and walked away.

Location: Shower and room

BARRY J. NIDORF (Sylmar) (2007, Age 16, 3 incidents)

PERPETRATOR 1: Plaintiff woke up sick, couldn't attend class. Requested to see nurse. After taking medicine, plaintiff passed out and woke in his room with perpetrator standing there looking at him. Upon going to restroom, plaintiff noticed yellow/clear mucus (sperm) on back of leg and near genitals area. Plaintiff believes perpetrator ejaculated on plaintiff's genitals and masturbated while plaintiff was unconscious. Perpetrator told plaintiff "everything we talk about is between us and I don't want you to get more time so don't say anything." Perpetrator also took pictures of plaintiff while showering.

PERPETRATOR 2: Gave plaintiff the medicine that made plaintiff drowsy and pass out. Made sexual comments toward plaintiff. Plaintiff believes Perpetrator 2 knew what Perpetrator 1 was going to do.

Location: Room and shower

CAMP MICHAEL SMITH (July 2007, Age 16, 1 incident)

Plaintiff woke to perpetrator being in room with weird pain/burning sensation. Plaintiff believes perpetrator gave him oral copulation.

Perpetrator told plaintiff to shower alone and humiliated plaintiff in front of other inmates by making plaintiff stay naked for long periods. Multiple sexual comments made.

On another occasion, perpetrator entered shower while plaintiff was showering and bathed plaintiff. Plaintiff felt it was complete mockery.

Location: Shower'''
    },
    # File 2: De La Mora, Alan (just read)
    {
        'Plaintiff Full Name': 'Alan De la Mora',
        'Completeness': '''Mostly Complete (80%)

✅ Personal Information: Complete
✅ Education History: Complete (no college)
✅ Employment History: Complete
⚠️ Probation History: Sparse (only 1 facility)
✅ Perpetrator Identification: Complete
✅ Abuse Details: Complete (over 15 incidents)
❌ Knowledge/Complaints: No complaints made
✅ Damages/Injuries: Complete
✅ Medical Treatment: Complete
⚠️ Loss of Earnings: Claims made, amounts TBD
✅ Other Damages: Complete
❌ Documents: None available''',
        'Perpetrator Details': '''BARRY J. NIDORF JUVENILE HALL (2012-2013)
Number: 1 perpetrator
Name: Rafael (went by "Rafa")
Description: Male, Hispanic, tall, dark hair, teacher's assistant/staff member''',
        'Abuse Details': '''BARRY J. NIDORF JUVENILE HALL (2012-2013, Age 15, over 15 incidents)

Every time plaintiff was in school, Rafa would sit next to him and start touching him around his thigh area while making sexualized comments. He would then proceed to fondle with plaintiff's penis under the clothing. He would rub his genitals against plaintiff's butt.

Rafa would take the plaintiff to the restroom and would digitally penetrate him; he would also use objects on the plaintiff to penetrate him.

Rafa would also show the plaintiff pornography and would masturbate in front of him.

Perpetrator gave plaintiff drugs in the facility in exchange for sexual favors, leading to plaintiff's drug addiction. The incident made plaintiff question his sexuality which led to further confusion, depression, anxiety and other mental issues.

Location: Restroom near the school area of the facility'''
    }
]

# Create DataFrame
df = pd.DataFrame(data_list)

# Save to Excel
output_file = 'Sample_Batch_6_Files.xlsx'
df.to_excel(output_file, index=False, engine='openpyxl')

# Format the Excel file
wb = load_workbook(output_file)
ws = wb.active

# Set column widths
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 80

# Enable text wrapping and top alignment
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')

# Set tall row heights for all data rows
for row_num in range(2, ws.max_row + 1):
    ws.row_dimensions[row_num].height = 800

wb.save(output_file)

print(f"✓ Sample batch Excel created: {output_file}")
print(f"  Files processed: 2 (Rodriguez + De La Mora)")
print(f"  Note: Processing remaining 4 files...")

